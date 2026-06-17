#!/usr/bin/env python3
"""
Transmittal Register Processor
Merges daily Aconex ExportMailAll with Master Submittal Log format.
"""

import re
import sys
import os
from datetime import datetime, timedelta
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import (PatternFill, Font, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter
import copy

# ─────────────────────────────────────────────────────────────────────────────
# Constants
# ─────────────────────────────────────────────────────────────────────────────

# Map short doc-type codes (from doc reference) to the Type codes used in Col H
DOC_TYPE_MAP = {
    'DRW': 'DRW', 'CAL': 'CAL', 'RPT': 'RPT', 'MTS': 'MAT', 'MAT': 'MAT',
    'ITP': 'ITP', 'MST': 'MST', 'PLN': 'PLN', 'PRD': 'PRD', 'PQN': 'PQN',
    'SPC': 'SPC', 'MAN': 'MAN', 'OTH': 'OTH', 'DES': 'DES', 'PGM': 'OTH',
    'IRE': 'OTH', 'TMP': 'PLN', 'SCH': 'OTH',
}

# Contractor org prefixes → outgoing submittals
CONTRACTOR_ORGS = {'MAPA-LIMAK-CRRC', 'GUNAL'}

# Engineer org prefix → responses
ENGINEER_ORGS_PREFIX = ('PAJV',)

# Full name → initials lookup for Lead Reviewer (Col L)
NAME_INITIALS = {
    'arunkumar ramakrishnan': 'AKR', 'mahmoud saad': 'MS', 'darren de silva': 'DAD',
    'daniel maddy': 'DM', 'yazan safieh': 'YS', 'rohit tharakan': 'RT',
    'lynneth santos': 'LS', 'ashok kumar muthuswamy': 'AKM', 'shahab ahmad khan': 'SAK',
    'ismaiel salama': 'IS', 'ahmad chikh bakri': 'ACB', 'mike lewey': 'ML',
    'amit kumar': 'AK', 'mazar syed': 'MSD', 'mohamed yahia': 'MY',
    'oladele kupolati': 'OK', 'john prabhu': 'JP', 'evripidis zampiras': 'EZ',
    'kaddour lardjani': 'KLI', 'stephen beadle': 'SAB', 'shaji aboobaker': 'SA',
    'alexander castellanos': 'AC', 'khaleel hamdan': 'KMH', 'reyad mahmoud': 'RM',
    'allan halferty': 'AH', 'fareed siddiqi': 'FS', 'loren leiski': 'LL',
    'vinod kalwaghe': 'VAK', 'ganesh chandra padhy': 'GCP', 'sherif zayed': 'SZ',
    'mohsin patel': 'MP', 'nageswara rao ampolu': 'NRA', 'christopher harvey': 'CH',
    'ezzeldin khattab': 'EK', 'suresh shetty': 'SSY', 'ahmed ibrahim': 'AI',
    'siddarth jayaraman': 'SRJ', 'alister white': 'AW', 'james varley': 'JV',
    'nuri suslu': 'NSU', 'sandeep jha': 'SJ', 'saipulezam ahmadnasir': 'SEN',
    'mohamed fahim abdelhady': 'MFA', 'dilip raju sarasa': 'DRS',
    'kent chin chang': 'KCC', 'sandip ghumde': 'SG', 'jason moore': 'JDM',
    'voon lee': 'VHL', 'fareed salih': 'FSL', 'reginald hingston': 'RH',
    'vishnu sadanand': 'VS', 'cristian bostan': 'CB', 'david song': 'DS',
    'roopa shah': 'RS', 'ahmed abdelhafez': 'AAH', 'andreas tauschinger': 'AT',
    'basil issac': 'BI', 'ahmed amin': 'AA', 'barkat ahmed': 'BA',
    'ade ogunsola': 'AO', 'ryan lamban': 'RL', 'naresh atmaram achpalia': 'NAA',
    'anna zouzoulas': 'AZ', 'greg watt': 'GW', 'lekha kumar': 'LSK',
    'yogesh kulkarni': 'YK',
}


def _build_name_aliases() -> dict:
    """Add aliases for 3+ word names where a middle/last name is dropped."""
    aliases = dict(NAME_INITIALS)
    for key, initials in NAME_INITIALS.items():
        words = key.split(' ')
        if len(words) >= 3:
            aliases.setdefault(f'{words[0]} {words[-1]}', initials)   # first + surname
            aliases.setdefault(f'{words[0]} {words[1]}', initials)    # first + middle
    return aliases


NAME_INITIALS_ALIASED = _build_name_aliases()


def lead_initials(name: str) -> str:
    """Map a full Lead name to its initials; return original name if no match."""
    if not name:
        return ''
    key = re.sub(r'\s+', ' ', name.strip().lower())
    return NAME_INITIALS_ALIASED.get(key, name)


# ─────────────────────────────────────────────────────────────────────────────
# Parsing helpers
# ─────────────────────────────────────────────────────────────────────────────

DOC_REF_RE = re.compile(
    r'[A-Z]{2,10}-[A-Z0-9]{2,10}-[A-Z]{2,5}-[A-Z]{2,5}(?:-[A-Z]{2,5})*-\d{4,6}'
    r'(?:\s+to\s+\d+)?'
)
DOT_REF_RE = re.compile(
    r'\b[A-Z]{2,4}\.[A-Z]{2,4}\.[A-Z]{2,4}\d{0,4}(?:\.[A-Z0-9]{1,5}){1,5}\b'
)
REV_RE = re.compile(r'_Rev[.\s]([A-Z0-9]+)', re.IGNORECASE)
DCP_RE = re.compile(r'\bDCP\s*(\d+)\b', re.IGNORECASE)


def extract_doc_refs(subject: str) -> tuple[list[str], bool]:
    """
    Extract doc refs from subject.
    Returns (refs, is_dot_format).
    """
    dash_refs = DOC_REF_RE.findall(subject or '')
    if dash_refs:
        return dash_refs, False
    dot_refs = DOT_REF_RE.findall(subject or '')
    return dot_refs, bool(dot_refs)


def parse_doc_ref(doc_ref: str) -> dict:
    """
    Parse a doc ref like MLCC-S0681-CIV-VBR-PIR-DRW-13000
    Returns dict with keys: discipline, sub_discipline, type_code, type_h
    """
    parts = doc_ref.split('-')
    result = {'discipline': '', 'sub_discipline': '', 'type_code': '', 'type_h': ''}
    if len(parts) >= 4:
        result['discipline'] = parts[2]          # e.g. CIV
        result['sub_discipline'] = parts[3]      # e.g. VBR
    if len(parts) >= 6:
        # The type code is typically the second-to-last alpha segment
        for seg in reversed(parts):
            if seg.isalpha() and len(seg) == 3:
                tc = seg.upper()
                result['type_code'] = tc
                result['type_h'] = DOC_TYPE_MAP.get(tc, tc)
                break
    return result


def extract_revision(subject: str) -> str:
    """Return the value after '_Rev.' e.g. '_Rev.C01' → 'C01', '_Rev.01' → '01'."""
    m = REV_RE.search(subject or '')
    return m.group(1) if m else ''


def extract_dcp(subject: str) -> str:
    m = DCP_RE.search(subject or '')
    return f"DCP{m.group(1)}" if m else ''


def parse_date(val) -> datetime | None:
    if val is None:
        return None
    if isinstance(val, datetime):
        return val
    if isinstance(val, str):
        for fmt in ('%d/%m/%Y', '%Y-%m-%d', '%m/%d/%Y'):
            try:
                return datetime.strptime(val.strip(), fmt)
            except ValueError:
                pass
    return None


def format_date(val) -> str:
    d = parse_date(val)
    return d.strftime('%d/%m/%Y') if d else (str(val) if val else '')


# ─────────────────────────────────────────────────────────────────────────────
# Main processor
# ─────────────────────────────────────────────────────────────────────────────

def load_master_log(path: str) -> dict:
    """Load master log, return dict keyed by Aconex Reference (col B)."""
    wb = load_workbook(path)
    ws = wb['Register']

    # Find header row (look for 'Aconex Reference')
    header_row = None
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=15, values_only=True), 1):
        if any('Aconex' in str(c) for c in row if c):
            header_row = i
            break
    if not header_row:
        raise ValueError("Could not find header row in Master Submittal Log")

    records = {}
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        aconex_ref = row[1]  # Col B
        if aconex_ref:
            records[str(aconex_ref).strip()] = {
                'row': list(row),
                'aconex_ref': str(aconex_ref).strip(),
            }
    return records


def load_export_mail(path: str) -> list[dict]:
    """Load ExportMailAll, return list of dicts."""
    wb = load_workbook(path)
    ws = wb['Mail']
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    records = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        rec = dict(zip(headers, row))
        records.append(rec)
    return records


def _first_after_colon(s: str) -> str:
    """Return first name after the first ':' in a string."""
    if not s:
        return ''
    idx = s.find(':')
    return s[idx + 1:].split(',')[0].strip() if idx != -1 else ''


def _norm_subject(s: str) -> str:
    """Strip leading Re: prefixes, collapse whitespace, lowercase for matching."""
    s = re.sub(r'^(Re:\s*)+', '', s or '', flags=re.IGNORECASE).strip().lower()
    return re.sub(r'\s+', ' ', s)


def build_intm_lead_map(mail_records: list[dict]) -> dict:
    """
    Map normalised subject → lead name.
    Source: INTM rows, lead = first name after ':' in their Recipients (col G).
    """
    lead_map = {}
    for rec in mail_records:
        if rec.get('Type') == 'Internal Memorandum':
            key = _norm_subject(rec.get('Subject', ''))[:120]
            lead = _first_after_colon(str(rec.get('Recipients') or ''))
            if key and lead:
                lead_map[key] = lead
    return lead_map


def process_transmittals(mail_records: list[dict],
                         existing_records: dict) -> list[dict]:
    """
    Build output rows.
    - Outgoing MLCC/GUNAL transmittals → new submittal rows
    - PAJV transmittals (engineer responses) are NOT given their own row.
      Instead, the PAJV reference (O) and response date (P) are matched
      by Submittal Item Description (col E) back onto the originating
      MLCC/GUNAL row — either a new row created today, or an existing
      master-log row.
    - Lead (col L): matched from INTM Recipients (col G) by subject
    Returns list of row dicts mapping to master log columns.
    """
    intm_lead_map = build_intm_lead_map(mail_records)

    # normalised description (col E) → aconex ref, for rows already in the master log
    master_ref_by_desc = {}
    for ref, info in existing_records.items():
        e = info['row'][4] if len(info['row']) > 4 else None
        if e:
            master_ref_by_desc[_norm_subject(str(e))] = ref

    new_rows: dict[str, dict] = {}     # mail_no → row, for today's MLCC/GUNAL transmittals
    update_rows: dict[str, dict] = {}  # aconex_ref → row, for existing master rows being updated
    unmatched_responses = []

    for rec in mail_records:
        if rec.get('Type') != 'Transmittal':
            continue
        org = str(rec.get('From Organization') or '')
        mail_no = str(rec.get('Mail No') or '').strip()
        is_contractor = any(org.startswith(c) for c in CONTRACTOR_ORGS) or \
                        (mail_no.startswith('MLCC-') or mail_no.startswith('GUNAL-'))
        if not is_contractor:
            continue

        subj = str(rec.get('Subject') or '').strip()
        date_val = parse_date(rec.get('Date'))

        doc_refs, is_dot_ref = extract_doc_refs(subj)
        first_ref = doc_refs[0] if doc_refs else ''
        parsed = {} if is_dot_ref else (parse_doc_ref(first_ref) if first_ref else {})

        current_rev = extract_revision(subj)
        dcp = extract_dcp(subj)
        lead = lead_initials(intm_lead_map.get(_norm_subject(subj)[:120], ''))

        row = {
            'A': '', 'B': mail_no, 'C': date_val,
            'D': first_ref if first_ref else '', 'E': subj,
            'F': current_rev, 'G': '',
            'H': parsed.get('type_h', ''), 'I': parsed.get('discipline', ''),
            'J': parsed.get('sub_discipline', ''),
            'K': len(doc_refs) if doc_refs else 1,
            'L': lead, 'M': dcp,
            'N': '', 'O': '', 'P': '', 'Q': '', 'R': '', 'S': '', 'T': '',
            'U': '', 'V': '', 'W': '', 'X': '', 'Y': '', 'Z': '',
            '_direction': 'outgoing',
            '_exists': mail_no in existing_records,
        }
        new_rows[mail_no] = row

    # normalised description (col E) → mail_no, for today's new rows (priority over master log)
    new_desc_to_mailno = {_norm_subject(r['E']): mail_no for mail_no, r in new_rows.items() if r['E']}

    for rec in mail_records:
        if rec.get('Type') != 'Transmittal':
            continue
        mail_no = str(rec.get('Mail No') or '').strip()
        if not mail_no.startswith('PAJV-TRANSMIT'):
            continue

        subj = str(rec.get('Subject') or '').strip()
        date_val = parse_date(rec.get('Date'))
        desc_key = _norm_subject(subj)
        doc_refs, _ = extract_doc_refs(subj)
        ref = doc_refs[0] if doc_refs else ''

        if desc_key in new_desc_to_mailno:
            target = new_rows[new_desc_to_mailno[desc_key]]
            target['O'] = mail_no
            target['P'] = date_val
        elif desc_key in master_ref_by_desc:
            aconex_ref = master_ref_by_desc[desc_key]
            if aconex_ref not in update_rows:
                master_row = existing_records[aconex_ref]['row']
                cols = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'
                update_rows[aconex_ref] = {
                    c: master_row[i] if i < len(master_row) else ''
                    for i, c in enumerate(cols)
                }
                update_rows[aconex_ref]['_direction'] = 'updated'
                update_rows[aconex_ref]['_exists'] = True
            update_rows[aconex_ref]['O'] = mail_no
            update_rows[aconex_ref]['P'] = date_val
        else:
            unmatched_responses.append({
                'A': '', 'B': '', 'C': '', 'D': ref, 'E': subj,
                'F': '', 'G': '', 'H': '', 'I': '', 'J': '', 'K': '',
                'L': '', 'M': '', 'N': '', 'O': mail_no, 'P': date_val,
                'Q': '', 'R': '', 'S': '', 'T': '', 'U': '', 'V': '',
                'W': '', 'X': '', 'Y': '',
                'Z': 'No matching MLCC/GUNAL submittal found for this response',
                '_direction': 'unmatched', '_exists': False,
            })

    return list(new_rows.values()) + list(update_rows.values()) + unmatched_responses


def write_output_excel(rows: list[dict], output_path: str,
                       template_path: str | None = None):
    """Write output Excel file formatted like the Master Submittal Log."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Register'

    # ── Header block ──────────────────────────────────────────────────────────
    HEADERS = [
        'Draft', 'Aconex Reference', 'Submittal Date',
        'Document Reference No./Package no.', 'Submittal Item Description',
        'Current Rev', 'Rev.', 'Type', 'Discipline', 'Sub Discipline',
        'No of Items', 'Lead', 'DCP',
        'Due date of Response by the Engineer (+21 days from Submittal date)',
        'Response Ref.', 'Date Responded by the Engineer',
        'Response Status, days', 'Days Overdue', 'Review Status',
        'Actual Status (Latest Rev)',
        'Contractor Response Due (+14 Days from Response Date)',
        'Contractor Date Responded', 'Contractor Days Overdue',
        'Contractor Response Status Days', 'Data Date', 'Remarks',
    ]

    # Title rows
    ws.merge_cells('A1:Z1')
    ws['A1'] = 'TRANSMITTAL REGISTER'
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')

    ws.merge_cells('A2:Z2')
    ws['A2'] = f'Generated: {datetime.now().strftime("%d/%m/%Y %H:%M")}'
    ws['A2'].alignment = Alignment(horizontal='center')

    # Blank rows 3-7 (match master log structure)
    for r in range(3, 8):
        ws.append([''] * 26)

    # Header row 8
    header_fill = PatternFill('solid', fgColor='1F4E79')
    header_font = Font(bold=True, color='FFFFFF', size=9)
    for col_idx, h in enumerate(HEADERS, 1):
        cell = ws.cell(row=8, column=col_idx, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center',
                                   wrap_text=True)

    ws.row_dimensions[8].height = 45

    # ── Data rows ─────────────────────────────────────────────────────────────
    COL_KEYS = list('ABCDEFGHIJKLMNOPQRSTUVWXYZ')

    alt_fill_new = PatternFill('solid', fgColor='EBF3FB')
    alt_fill_exists = PatternFill('solid', fgColor='FFF2CC')
    thin = Side(style='thin', color='CCCCCC')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    data_font = Font(size=9)
    data_font_exists = Font(size=9, italic=True, color='808080')

    outgoing_fill   = PatternFill('solid', fgColor='E2EFDA')   # new submittal rows
    updated_fill    = PatternFill('solid', fgColor='DDEBF7')   # existing rows getting a response today
    unmatched_fill  = PatternFill('solid', fgColor='FCE4D6')   # orphaned PAJV responses

    FILL_BY_DIRECTION = {
        'outgoing': outgoing_fill,
        'updated': updated_fill,
        'unmatched': unmatched_fill,
    }

    new_count       = len([r for r in rows if r['_direction'] == 'outgoing' and not r['_exists']])
    exists_count    = len([r for r in rows if r['_direction'] == 'outgoing' and r['_exists']])
    updated_count   = len([r for r in rows if r['_direction'] == 'updated'])
    unmatched_count = len([r for r in rows if r['_direction'] == 'unmatched'])

    print(f"New submittal rows:                {new_count}")
    print(f"Already in master log (flagged):   {exists_count}")
    print(f"Existing rows updated with response: {updated_count}")
    print(f"Unmatched PAJV responses:          {unmatched_count}")

    row_num = 9
    for row in rows:
        fill = FILL_BY_DIRECTION.get(row['_direction'], outgoing_fill)
        if row['_exists'] and row['_direction'] == 'outgoing':
            fill = alt_fill_exists

        for col_idx, key in enumerate(COL_KEYS, 1):
            val = row.get(key, '')
            cell = ws.cell(row=row_num, column=col_idx, value=val)
            cell.fill = fill
            cell.font = data_font_exists if row['_exists'] else data_font
            cell.border = border
            if key == 'E':
                cell.alignment = Alignment(wrap_text=True)
            elif key in ('C', 'N', 'P', 'U', 'Y'):
                if isinstance(val, datetime):
                    cell.number_format = 'DD/MM/YYYY'
            else:
                cell.alignment = Alignment(horizontal='left', vertical='center')

        if row['_exists'] and row['_direction'] == 'outgoing':
            ws.cell(row=row_num, column=26, value='[Already in master log]')

        row_num += 1

    # ── Column widths ─────────────────────────────────────────────────────────
    COL_WIDTHS = {
        1: 8, 2: 22, 3: 14, 4: 38, 5: 55, 6: 10, 7: 8, 8: 8, 9: 10,
        10: 12, 11: 8, 12: 10, 13: 8, 14: 18, 15: 22, 16: 14, 17: 14,
        18: 12, 19: 14, 20: 14, 21: 14, 22: 14, 23: 14, 24: 14, 25: 12, 26: 30,
    }
    for col_idx, width in COL_WIDTHS.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Freeze panes at row 9, col C
    ws.freeze_panes = 'C9'

    # ── Summary sheet ─────────────────────────────────────────────────────────
    ws_sum = wb.create_sheet('Summary')
    ws_sum['A1'] = 'Processing Summary'
    ws_sum['A1'].font = Font(bold=True, size=12)
    summary_data = [
        ['Total rows produced', len(rows)],
        ['New outgoing submittals', new_count],
        ['Already in master log', exists_count],
        ['Existing rows updated with PAJV response', updated_count],
        ['Unmatched PAJV responses (no submittal found)', unmatched_count],
        ['Generated', datetime.now().strftime('%d/%m/%Y %H:%M')],
    ]
    for r_idx, (label, val) in enumerate(summary_data, 3):
        ws_sum.cell(row=r_idx, column=1, value=label)
        ws_sum.cell(row=r_idx, column=2, value=val)
    ws_sum.column_dimensions['A'].width = 35
    ws_sum.column_dimensions['B'].width = 20

    wb.save(output_path)
    print(f"\nOutput saved to: {output_path}")


# ─────────────────────────────────────────────────────────────────────────────
# Entry point
# ─────────────────────────────────────────────────────────────────────────────

def run(master_path: str, export_mail_path: str, output_path: str):
    print("Loading master log...")
    existing = load_master_log(master_path)
    print(f"  Loaded {len(existing)} existing records")

    print("Loading export mail...")
    mail_records = load_export_mail(export_mail_path)
    print(f"  Loaded {len(mail_records)} mail records")

    print("Processing transmittals...")
    rows = process_transmittals(mail_records, existing)
    print(f"  Built {len(rows)} output rows")

    print("Writing output Excel...")
    write_output_excel(rows, output_path)


if __name__ == '__main__':
    master = sys.argv[1] if len(sys.argv) > 1 else \
        '/root/.claude/uploads/8c61c79c-4d5d-56a0-a9bd-29d06cd54f04/e07e6045-Master_Submittal_Log_10062026.xlsx'
    export = sys.argv[2] if len(sys.argv) > 2 else \
        '/root/.claude/uploads/8c61c79c-4d5d-56a0-a9bd-29d06cd54f04/dfd6085b-ExportMailAll20260613_0837.xlsx'
    output = sys.argv[3] if len(sys.argv) > 3 else \
        '/home/user/Coursiv-training-/output_transmittal_register.xlsx'

    run(master, export, output)
